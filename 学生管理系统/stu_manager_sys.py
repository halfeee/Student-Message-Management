# -*- coding: utf-8 -*-
from flask import Flask, render_template, request,session
import cgi
import codecs
import sys
import sqlite3
import os
import random
import openpyxl

app = Flask(__name__)
app.secret_key="123456789"

#登录进入初始界面
@app.route('/',methods=['GET'])
def index():

    return render_template('login.html');

#返回按钮：
@app.route('/back',methods=['GET'])
def back():
    if(session['login_type']=='管理人员'):
        return render_template('manager.html', account=session['account'], name=session['name'], sex=session['sex']);
    elif(session['login_type']=='教师'):
        return render_template('teacher.html', account=session['account'], name=session['name'], sex=session['sex'],major=session['major']);
    elif (session['login_type'] == '学生'):
        return render_template('student.html',account=session['account'] , name=session['name'],sex = session['sex'],day = session['day'],class_name = session['class'] );
    else:
        return render_template('login.html');




#登录信息处理
@app.route('/login',methods=['POST','GET'])
def login():
    account = "";
    password = "";
    if 'account' not in request.args:
        account = "";
    else:
        account = request.args['account'];
    if 'password' not in request.args:
        password = "1";
    else:
        password = request.args['password'];

#检测是否为管理员
    sql = "select * from managers where id ='"+account+"'and password='"+password+"'";
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(sql)
    result = cur.fetchall()  # 获得查询的结果
    cur.close()
    print(sql)
    print(result.__len__());
    if(result.__len__()!=0):
        session['account']=result[0][0];
        session['name'] = result[0][1];
        session['sex'] = result[0][2];
        session['password'] = result[0][3];
        session['login_type'] = "管理人员";
        return render_template('manager.html',account=session['account'],name=session['name'],sex=session['sex'],password = session['password']);
#检测是否为学生
    sql = "select * from student where id ='" + account + "'and password='" + password + "'";
    cur = conn.cursor()
    cur.execute(sql)
    result = cur.fetchall()  # 获得查询的结果
    #cur.close()
    print(sql)
    print(result.__len__());
    if (result.__len__() != 0):
        session['account'] = result[0][0];
        session['name'] = result[0][1];
        session['sex'] = result[0][2];
        session['day'] = result[0][3];
        #session['class'] = result[0][4];
        session['password'] = result[0][5];
        session['login_type'] = "学生";
        sql = "select name from class where id = '%s'"%(result[0][4]);
        cur.execute(sql);
        session['class'] = cur.fetchall()[0][0];
        return render_template('student.html',account=session['account'] , name=session['name'],sex = session['sex'],day = session['day'],class_name = session['class'],password = session['password'] );
#检测是否为教师
    sql = "select * from teacher where id ='" + account + "'and password='" + password + "'";
    cur = conn.cursor()
    cur.execute(sql)
    result = cur.fetchall()  # 获得查询的结果

    print(sql)
    print(result.__len__());
    if (result.__len__() != 0):
        session['account'] = result[0][0];
        session['name'] = result[0][1];
        session['sex'] = result[0][2];
        session['major_id'] = result[0][3];
        session['password'] = result[0][4];
        session['login_type'] = "教师";
        cur.execute("select name from major where id = '%s'"%(result[0][3]));
        major_list = cur.fetchall();
        major = major_list[0][0]
        session['major'] = major;
        return render_template('teacher.html', account=session['account'],name = session['name'],sex = session['sex'],major = session['major'],password = session['password'] );

    return render_template('login.html', message="用户名或密码错误");

@app.route('/manager',methods=['GET'])
def to_change_manager_inf():
    if 'login_type' not in session:
        return render_template('login.html');
    return render_template('manager_chg_information.html',account=session['account'],name=session['name'],sex=session['sex'],password = session['password']);

@app.route('/stu_delete',methods=['GET'])
def stu_delete():
    if 'login_type' not in session:
        return render_template('login.html');
    account = session['account'];
    name = session['name'];
    sex = session['sex'];
    stu_id = request.args['stu_id'];
    print(account+" "+name+" "+sex+" "+stu_id);
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    #更改数据库数据
    cur = conn.cursor()
    cur.execute("DELETE FROM student where id ='"+stu_id+"'");
    conn.commit();
    #查询数据
    cur = conn.cursor()
    cur.execute("select student.id,student.name,student.sex,student.birthday,class.name,password from student,class where class.id = student.class_num;");
    result = cur.fetchall()  # 获得查询的结果
    cur.close()
    return render_template('manager_stu_list.html',account=account,name=name,sex=sex,stu_table=result);

@app.route('/m_student',methods=['GET'])
def m_student():
    if 'login_type' not in session:
        return render_template('login.html');
    id = session['account'];
    name = session['name'];
    sex = session['sex'];
    #数据库查询学生表
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute("select student.id,student.name,student.sex,student.birthday,class.name,password from student,class where class.id = student.class_num;");
    result = cur.fetchall()  # 获得查询的结果
    cur.close()
    return render_template('manager_stu_list.html',account=session['account'],name=session['name'],sex=session['sex'],stu_table=result);

@app.route('/select_student',methods=['GET'])
def select_student():
    if 'login_type' not in session:
        return render_template('login.html');
    id = session['account'];
    name = session['name'];
    sex = session['sex'];
    stu_name = request.args['stu_name'];
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    sql = "select student.id,student.name,student.sex,student.birthday,class.name,password from student,class where class.id = student.class_num and student.name like '%%%s%%';"%(stu_name);
    cur.execute(sql)
    result = cur.fetchall()  # 获得查询的结果
    cur.close()
    return render_template('manager_stu_list.html', account=session['account'], name=session['name'],sex=session['sex'], stu_table=result);


@app.route('/manager_change',methods=['GET'])
def manager_change():
    if 'login_type' not in session:
        return render_template('login.html');
    account = request.args['new_account'];
    name = request.args['new_name'];
    password = request.args['new_password'];
    sex = request.args['new_sex'];
    sql = "update managers set name = '"+name+"' ,sex = '"+sex+"',password = '"+password+"'  where id='"+account+"'"
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(sql);
    conn.commit();
    cur.close();
    print(sql);
    return render_template('login.html');

@app.route('/to_stu_update',methods=['GET'])
def to_stu_update():
    if 'login_type' not in session:
        return render_template('login.html');
    stu_id = request.args['stu_id'];
    sql =  "select * from student where id = '"+stu_id+"'";
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(sql);
    result = cur.fetchall()  # 获得查询的结果
    result = result[0];
    cur.execute("select name from class");
    class_name = cur.fetchall();
    sql = "select name from class where id = '%s'"%(result[4]);
    cur.execute(sql);
    class_name_n = cur.fetchall()[0][0];
    print(result);
    return render_template('stu_update.html',result = result,class_name = class_name,class_name_n = class_name_n);

@app.route('/stu_update',methods=['GET'])
def stu_update():
    if 'login_type' not in session:
        return render_template('login.html');
    new_account = request.args['new_account'];
    new_name = request.args['new_name'];
    new_sex = request.args['new_sex'];
    new_date = request.args['new_date'];
    new_class = request.args['new_class'];
    new_password = request.args['new_password'];
    #修改数据库数据
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    #转换班级名-》班级号
    sql = "select id from class where name = '%s'"%(new_class);
    cur.execute(sql);
    new_class = cur.fetchall()[0][0];
    cur = conn.cursor();
    #组建新的更新语句
    sql = "update student set name='%s',sex='%s',birthday='%s',class_num='%s',password='%s' where id = '%s'" % (new_name, new_sex, new_date, new_class, new_password, new_account);
    cur.execute(sql);
    conn.commit();
    cur.close();
    #查询新的学生表
    cur = conn.cursor()
    cur.execute("select student.id,student.name,student.sex,student.birthday,class.name,password from student,class where class.id = student.class_num;");
    result = cur.fetchall()  # 获得查询的结果
    cur.close()
    if(session['login_type']=="管理人员"):
        return render_template('manager_stu_list.html', account=session['account'], name=session['name'],sex=session['sex'], stu_table=result);
    if(session['login_type']=="学生"):
        return render_template('login.html');


@app.route('/m_teacher',methods=['GET'])
def m_teacher():
    if 'login_type' not in session:
        return render_template('login.html');
    id = session['account'];
    name = session['name'];
    sex = session['sex'];
    # 数据库查询学生表
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute("select teacher.id,teacher.name,teacher.sex,major.name,major.department,teacher.password from teacher,major where teacher.major=major.id");
    result = cur.fetchall()  # 获得查询的结果
    cur.close()
    return render_template('manager_tea_list.html', account=session['account'], name=session['name'], sex=session['sex'], tea_table=result);

@app.route('/select_teacher',methods=['GET'])
def select_teacher():
    if 'login_type' not in session:
        return render_template('login.html');

    tea_name = request.args['tea_name'];
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    sql = "select teacher.id,teacher.name,teacher.sex,major.name,major.department,teacher.password from teacher,major where teacher.major=major.id and teacher.name like '%%%s%%';" % (tea_name);
    cur.execute(sql)
    result = cur.fetchall()  # 获得查询的结果
    cur.close()
    return render_template('manager_tea_list.html', account=session['account'], name=session['name'],sex=session['sex'], tea_table=result);


@app.route('/tea_delete',methods=['GET'])
def tea_delete():
    if 'login_type' not in session:
        return render_template('login.html');
    account = session['account'];
    name = session['name'];
    sex = session['sex'];
    tea_id = request.args['tea_id'];
    print(account+" "+name+" "+sex+" "+tea_id);
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    #更改数据库数据
    cur = conn.cursor()
    cur.execute("DELETE FROM teacher where id ='"+tea_id+"'");
    conn.commit();
    #查询数据
    cur = conn.cursor()
    cur.execute("select teacher.id,teacher.name,teacher.sex,major.name,major.department,teacher.password from teacher,major where teacher.major=major.id");
    result = cur.fetchall()  # 获得查询的结果
    cur.close()
    return render_template('manager_tea_list.html',account=account,name=name,sex=sex,tea_table=result);

@app.route('/to_tea_update',methods=['GET'])
def to_tea_update():
    if 'login_type' not in session:
        return render_template('login.html');
    tea_id = request.args['tea_id'];
    print('========================================'+tea_id);
    sql =  "select teacher.id,teacher.name,teacher.sex,major.name,teacher.password from teacher,major where teacher.major=major.id and teacher.id=%s"%(tea_id);
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(sql);
    result = cur.fetchall()  # 获得查询的结果
    result = result[0];
    print(result);
    return render_template('tea_update.html',result = result);
@app.route('/to_my_course',methods=['GET'])
def to_my_course():
    if 'login_type' not in session:
        return render_template('login.html');
    stu_id = request.args['stu_id'];
    sql = "select course.id,course.name,teacher.name,grade.grade from teacher,course,grade,student where teacher.id = course.teacher_id and grade.student_id = student.id and grade.class_id = course.id and student.id = '%s';"%(stu_id);
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(sql);
    result = cur.fetchall()  # 获得查询的结果
    return render_template('my_course.html',result = result);
@app.route('/to_add_my_course',methods=['GET'])
def to_add_my_course():
    if 'login_type' not in session:
        return render_template('login.html');
    stu_id = session['account'];
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    sql = "select grade.class_id from grade where student_id = '%s'"%(stu_id);
    print(sql);
    cur.execute(sql);
    result = cur.fetchall();
    list = "("
    print("result"+str(result));
    for id in result:
        list = list+"'"+id[0]+"',"
    list = list+"'XXXXX')"
    print(list);
    sql = "select course.*,teacher.name from course,teacher where course.teacher_id = teacher.id and course.id not in"+list;
    cur.execute(sql);
    result = cur.fetchall();
    return render_template('add_my_course.html',table=result);

@app.route('/add_stu_course',methods=['GET'])
def add_stu_course():
    if 'login_type' not in session:
        return render_template('login.html');
    stu_id = session['account'];
    course_id = request.args['course_id'];
    id = int(random.uniform(0,1000000));
    sql = "INSERT INTO grade VALUES ('%s','%s','%s','0')"%(id,stu_id,course_id);
    database = "database\Stu_Message_manger.db";
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    print("sql1=========="+sql)
    cur.execute(sql);
    conn.commit();
    sql = "select course.stu_num from course where course.id ='%s' "%(course_id);
    cur.execute(sql);
    num =  cur.fetchall()[0][0];
    num = int(num);
    num = num+1;
    sql = "update course set stu_num = '%s' where id = '%s'"%(num,course_id);
    cur.execute(sql);
    conn.commit();
    print("sql2==========" + sql)
    sql = "select course.id,course.name,teacher.name,grade.grade from teacher,course,grade,student where teacher.id = course.teacher_id and grade.student_id = student.id and grade.class_id = course.id and student.id = '%s';" %(stu_id);
    cur.execute(sql);
    result = cur.fetchall()  # 获得查询的结果
    return render_template('my_course.html', result=result);


@app.route('/tea_update',methods=['GET'])
def tea_update():
    if 'login_type' not in session:
        return render_template('login.html');
    new_account = request.args['new_account'];
    new_name = request.args['new_name'];
    new_sex = request.args['new_sex'];
    new_major = request.args['new_major'];
    new_password = request.args['new_password'];
    #连接数据库
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)

    #转化专业号
    if(new_major=="计算机科学与技术"):
        new_major = "01"
    if (new_major == "数字媒体"):
        new_major = "02"
    if (new_major == "电子技术应用"):
        new_major = "03"
    if (new_major == "软件工程"):
        new_major = "04"
    if (new_major == "网络工程"):
        new_major = "05"
    if (new_major == "智能科学与技术"):
        new_major = "06"
    if (new_major == "信息安全"):
        new_major = "07"
    if (new_major == "电子信息工程"):
        new_major = "08"

    #修改数据库数据
    sql = "update teacher set name='%s',sex='%s',major='%s',password='%s' where id = '%s'"%(new_name,new_sex,new_major , new_password,new_account);
    print(sql);
    cur = conn.cursor()
    cur.execute(sql);
    conn.commit();
    cur.close();

    #查询新的教师表
    if session['login_type'] == '管理人员':
        cur = conn.cursor()
        cur.execute("select teacher.id,teacher.name,teacher.sex,major.name,major.department,teacher.password from teacher,major where teacher.major=major.id");
        result = cur.fetchall()  # 获得查询的结果
        cur.close()
        return render_template('manager_tea_list.html', account=session['account'], name=session['name'],sex=session['sex'], tea_table=result);
    if session['login_type']=='教师':
        return render_template('login.html');
@app.route('/m_class',methods=['GET'])
def m_class():
    if 'login_type' not in session:
        return render_template('login.html');
    id = session['account'];
    name = session['name'];
    sex = session['sex'];
    # 数据库查询学生表
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute("select course.id,course.name,teacher.name,course.max_num,course.stu_num,course.time,course.credit,course.start_term,course.class_type from course,teacher where course.teacher_id = teacher.id");
    result = cur.fetchall()  # 获得查询的结果
    cur.close()
    return render_template('manager_class_list.html', account=session['account'], name=session['name'], sex=session['sex'], class_table=result);

@app.route('/class_delete',methods=['GET'])
def class_delete():
    if 'login_type' not in session:
        return render_template('login.html');
    account = session['account'];
    name = session['name'];
    sex = session['sex'];
    class_id = request.args['class_id'];
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    #更改数据库数据
    cur = conn.cursor()
    cur.execute("DELETE FROM course where id ='"+class_id+"'");
    conn.commit();
    #查询数据
    cur = conn.cursor()
    cur.execute("select course.id,course.name,teacher.name,course.max_num,course.stu_num,course.time,course.credit,course.start_term,course.class_type from course,teacher where course.teacher_id = teacher.id");
    result = cur.fetchall()  # 获得查询的结果
    cur.close()
    return render_template('manager_class_list.html', account=session['account'], name=session['name'], sex=session['sex'], class_table=result);


@app.route('/to_class_update',methods=['GET'])
def to_class_update():
    if 'login_type' not in session:
        return render_template('login.html');
    class_id = request.args['class_id'];
    print("classid                     :" +class_id);
    sql =  "select course.id,course.name,teacher.name,course.max_num,course.stu_num,course.time,course.credit,course.start_term,course.class_type from course,teacher where course.teacher_id = teacher.id and course.id='%s'"%(class_id)
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(sql);
    result = cur.fetchall()  # 获得查询的结果
    result = result[0];
    cur.execute("select name from teacher");
    teacher_name_list = cur.fetchall();
    return render_template('class_update.html',result = result,teacher_name_list = teacher_name_list);

@app.route('/class_update',methods=['GET'])
def class_update():
    if 'login_type' not in session:
        return render_template('login.html');
    new_account = request.args['new_account'];
    new_name = request.args['new_name'];
    new_teacher = request.args['new_teacher'];
    new_max_num = request.args['new_max_num'];
    new_time = request.args['new_time'];
    new_credit = request.args['new_credit'];
    new_start_term = request.args['new_start_term'];
    new_type = request.args['new_type'];
    #查询老师编号
    database = "database\Stu_Message_manger.db";
    conn = sqlite3.connect(database);
    cur = conn.cursor()
    sql = "select teacher.id from teacher where teacher.name = '%s'"%(new_teacher);
    cur.execute(sql);
    result = cur.fetchall();
    new_teacher = result[0][0];
    print("new teacher===========================================================:"+new_teacher);
    #修改数据库数据
    database = "database\Stu_Message_manger.db"
    sql = "update course set name = '%s',teacher_id = '%s' ,max_num = '%s',time = '%s',credit = '%s',start_term = '%s',class_type = '%s' where course.id = '%s'"%(new_name,new_teacher,new_max_num,new_time,new_credit,new_start_term,new_type,new_account);
    print(sql);
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(sql);
    conn.commit();
    cur.close();

    #查询新的课程表
    cur = conn.cursor()
    cur.execute("select course.id,course.name,teacher.name,course.max_num,course.stu_num,course.time,course.credit,course.start_term,course.class_type from course,teacher where course.teacher_id = teacher.id");
    result = cur.fetchall()  # 获得查询的结果
    cur.close()
    return render_template('manager_class_list.html', account=session['account'], name=session['name'], sex=session['sex'], class_table=result);
@app.route('/to_add_student',methods=['GET'])
def to_add_student():
    if 'login_type' not in session:
        return render_template('login.html');

    sql = "select name from class"
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(sql);
    result = cur.fetchall()  # 获得查询的结果
    return render_template('add_student.html',class_name = result)
@app.route('/stu_add',methods=['GET'])
def stu_add():
    if 'login_type' not in session:
        return render_template('login.html');
    new_account = request.args['new_account'];
    new_name = request.args['new_name'];
    new_sex = request.args['new_sex'];
    new_date = request.args['new_date'];
    new_class = request.args['new_class'];
    new_password = request.args['new_password'];

    #class_name->class_id
    database = "database\Stu_Message_manger.db";
    conn = sqlite3.connect(database);
    cur = conn.cursor()
    sql = "select id from class where name = '%s'"%(new_class);
    cur.execute(sql);
    new_class = cur.fetchall()[0][0];
    sql = "INSERT INTO student VALUES ('%s', '%s', '%s','%s','%s','%s' )"%(new_account,new_name,new_sex,new_date,new_class,new_password)


    cur.execute(sql);
    conn.commit();
    cur.close();
    # 查询新的学生表
    cur = conn.cursor()
    cur.execute("select student.id,student.name,student.sex,student.birthday,class.name,password from student,class where class.id = student.class_num;");
    result = cur.fetchall()  # 获得查询的结果
    cur.close()
    return render_template('manager_stu_list.html', account=session['account'], name=session['name'],sex=session['sex'], stu_table=result);

@app.route('/to_add_teacher',methods=['GET'])
def to_add_teacher():
    if 'login_type' not in session:
        return render_template('login.html');
    return render_template('add_teacher.html')

@app.route('/tea_add',methods=['GET'])
def tea_add():
    if 'login_type' not in session:
        return render_template('login.html');
    new_account = request.args['new_account'];
    new_name = request.args['new_name'];
    new_sex = request.args['new_sex'];
    new_major = request.args['new_major'];
    new_password = request.args['new_password'];
    # 连接数据库
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)

    # 转化专业号
    if (new_major == "计算机科学与技术"):
        new_major = "01"
    if (new_major == "数字媒体"):
        new_major = "02"
    if (new_major == "电子技术应用"):
        new_major = "03"
    if (new_major == "软件工程"):
        new_major = "04"
    if (new_major == "网络工程"):
        new_major = "05"
    if (new_major == "智能科学与技术"):
        new_major = "06"
    if (new_major == "信息安全"):
        new_major = "07"
    if (new_major == "电子信息工程"):
        new_major = "08"
#修改数据库数据
    sql = "INSERT INTO teacher VALUES ('%s', '%s','%s','%s','%s' )"%(new_account,new_name,new_sex,new_major,new_password)
    print(sql);
    cur = conn.cursor()
    cur.execute(sql);
    conn.commit();
    cur.close();
#跟新表格
    cur = conn.cursor()
    cur.execute("select teacher.id,teacher.name,teacher.sex,major.name,major.department,teacher.password from teacher,major where teacher.major=major.id");
    result = cur.fetchall()  # 获得查询的结果
    cur.close()
    return render_template('manager_tea_list.html', account=session['account'], name=session['name'],sex=session['sex'], tea_table=result);

@app.route('/to_add_class',methods=['GET'])
def to_add_class():
    if 'login_type' not in session:
        return render_template('login.html');
    database = "database\Stu_Message_manger.db"
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute("select name from teacher");
    teacher_name_list = cur.fetchall();
    return render_template('add_class.html',teacher_name_list = teacher_name_list);
@app.route('/class_add',methods=['GET'])
def class_add():
    if 'login_type' not in session:
        return render_template('login.html');
    new_account = request.args['new_account'];
    new_name = request.args['new_name'];
    new_teacher = request.args['new_teacher'];
    new_max_num = request.args['new_max_num'];
    new_time = request.args['new_time'];
    new_credit = request.args['new_credit'];
    new_start_term = request.args['new_start_term'];
    new_type = request.args['new_type'];
    determine = "正常";
    if(new_account==""):
        determine="请补全信息"
    elif (new_name == ""):
        determine = "请补全信息"
    elif (new_teacher == ""):
        determine = "请补全信息"
    elif (new_max_num == ""):
        determine = "请补全信息"
    elif (new_credit == ""):
        determine = "请补全信息"
    elif (new_start_term == ""):
        determine = "请补全信息"
    elif (new_type == ""):
        determine = "请补全信息"
    elif(not is_number(new_max_num)):
        determine = "学生数量错误"
    elif (not is_number(new_time)):
        determine = "学时错误"
    elif(not is_number(new_credit)):
        determine = "学分错误"
    # 查询老师编号
    if(determine == "正常"):
        database = "database\Stu_Message_manger.db";
        conn = sqlite3.connect(database);
        cur = conn.cursor()
        sql = "select teacher.id from teacher where teacher.name = '%s'" % (new_teacher);
        cur.execute(sql);
        result = cur.fetchall();
        new_teacher = result[0][0];
        sql = "INSERT INTO course VALUES ('%s', '%s', '%s','%s','0','%s','%s','%s','%s')"%(new_account,new_name,new_teacher,new_max_num,new_time,new_credit,new_start_term,new_type);
        cur.execute(sql);
        conn.commit();
        cur.execute("select course.id,course.name,teacher.name,course.max_num,course.stu_num,course.time,course.credit,course.start_term,course.class_type from course,teacher where course.teacher_id = teacher.id");
        result = cur.fetchall()  # 获得查询的结果
        cur.close()
        return render_template('manager_class_list.html', account=session['account'], name=session['name'],sex=session['sex'], class_table=result);
    else:
        database = "database\Stu_Message_manger.db"
        conn = sqlite3.connect(database)
        cur = conn.cursor()
        cur.execute("select name from teacher");
        teacher_name_list = cur.fetchall();
        return render_template('add_class.html', teacher_name_list=teacher_name_list,message = determine);

@app.route('/to_tea_course',methods=['GET'])
def to_tea_class():
    if 'login_type' not in session:
        return render_template('login.html');
    tea_id = request.args['tea_id'];
    sql = "select course.id,course.name,teacher.name,course.max_num,course.stu_num,course.time,course.credit,course.start_term,course.class_type from course,teacher where course.teacher_id = teacher.id and teacher_id = '%s'"%(tea_id);
    database = "database\Stu_Message_manger.db";
    conn = sqlite3.connect(database);
    cur = conn.cursor();
    cur.execute(sql);
    result = cur.fetchall()
    return render_template('tea_course.html',class_table=result);

@app.route('/course_detail',methods=['GET'])
def class_detail():
    if 'login_type' not in session:
        return render_template('login.html');
    course_id = request.args['course_id'];
    sql = "select grade.id ,course.name,student.name,grade.grade from course,student,grade where grade.student_id = student.id and grade.class_id = course.id and course.id = '%s'"%(course_id);
    database = "database\Stu_Message_manger.db";
    conn = sqlite3.connect(database);
    cur = conn.cursor();
    cur.execute(sql);
    result = cur.fetchall()
    return render_template('course_grade.html', grade_table=result,course_id = course_id);


@app.route('/tea_change_grade',methods=['GET'])
def tea_change_grade():
    if 'login_type' not in session:
        return render_template('login.html');
    course_id = request.args['course_id'];
    grade_id = request.args['grade_id'];
    new_grade = request.args['new_grade'];
    print(" = = = == = = = = = = = = = = = = == = == == =  "+course_id+" "+grade_id+" "+new_grade);
    database = "database\Stu_Message_manger.db";
    conn = sqlite3.connect(database);
    cur = conn.cursor();
    cur.execute("update grade set grade = '%s' where id = '%s'"%(new_grade,grade_id));
    conn.commit();
    sql = "select grade.id ,course.name,student.name,grade.grade from course,student,grade where grade.student_id = student.id and grade.class_id = course.id and course.id = '%s'"%(course_id);
    cur.execute(sql);
    result = cur.fetchall()
    return render_template('course_grade.html', grade_table=result, course_id=course_id,message='修改完成');

@app.route('/add_grade',methods=['GET'])
def add_grade():
    course_id = request.args['course_id'];
    mytext = request.args['mytext'];
    print(course_id);#获取的课程号
    print("==============================",mytext);#获取的文件路径
    database = "database\Stu_Message_manger.db";
    conn = sqlite3.connect(database);
    cur = conn.cursor();
    wb = openpyxl.load_workbook(mytext)
    ws=wb.active;
    print(ws.max_row,ws.max_column);
    for i in range(1,ws.max_row+1):
        for j in ['A','B','C']:
            x = j+str(i);
            if j=='A':
                id = ws[x].value;
            if j=='C':
                grade =ws[x].value;
        sql = "update grade set grade = '%s' where student_id = '%s' and class_id = '%s'"%(grade,id,course_id)
        cur.execute(sql);
        conn.commit();


    sql = "select grade.id ,course.name,student.name,grade.grade from course,student,grade where grade.student_id = student.id and grade.class_id = course.id and course.id = '%s'" % (course_id);
    cur.execute(sql);
    result = cur.fetchall()
    return render_template('course_grade.html', grade_table=result, course_id=course_id, message='修改完成');




def is_number(str):
  try:
    # 因为使用float有一个例外是'NaN'
    if str=='NaN':
      return False
    float(str)
    return True
  except ValueError:
    return False

if __name__ == '__main__':
    app.run(debug=True)
